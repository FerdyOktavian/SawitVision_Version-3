"""
Layanan pembuatan laporan Excel SawitVision V3.

Perubahan utama V3:
- data pengguna memakai nomor telepon;
- tidak ada kolom email;
- tidak ada status verifikasi email;
- struktur laporan tetap mempertahankan ringkasan, grafik, dan detail prediksi.
"""

import io
import re
from datetime import date, datetime, time
from zoneinfo import ZoneInfo
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text


JAKARTA_TZ = ZoneInfo("Asia/Jakarta")


def _excel_safe_datetime(value):
    """Ubah datetime timezone-aware menjadi WIB tanpa tzinfo agar aman untuk Excel."""
    if not isinstance(value, datetime):
        return value

    if value.tzinfo is not None:
        value = value.astimezone(JAKARTA_TZ).replace(tzinfo=None)

    return value


CLASS_LABELS = {
    "belum_masak": "Belum Masak",
    "masak": "Masak",
    "terlalu_masak": "Terlalu Masak",
}

HEADER_FILL = PatternFill("solid", fgColor="2F7D32")
SUBHEADER_FILL = PatternFill("solid", fgColor="E8F1E5")
ACCENT_FILL = PatternFill("solid", fgColor="F4EAD8")
TITLE_FILL = PatternFill("solid", fgColor="24351F")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_\-]+", "_", value or "user")
    return cleaned.strip("_") or "user"


def build_period_label(
    start_date: Optional[date],
    end_date: Optional[date],
) -> str:
    if start_date and end_date:
        return f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    if start_date:
        return f"Mulai {start_date.strftime('%d/%m/%Y')}"
    if end_date:
        return f"Sampai {end_date.strftime('%d/%m/%Y')}"
    return "Semua periode"


def _start_datetime(value: Optional[date]) -> Optional[datetime]:
    return datetime.combine(value, time.min) if value else None


def _end_datetime(value: Optional[date]) -> Optional[datetime]:
    return datetime.combine(value, time.max) if value else None


def get_user_prediction_report_data(
    db,
    user_id: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> Dict[str, Any]:
    user_row = db.execute(
        text("""
            SELECT id, full_name, phone_number, created_at
            FROM users
            WHERE id = :user_id
            LIMIT 1
        """),
        {"user_id": user_id},
    ).fetchone()

    if not user_row:
        raise ValueError("Pengguna tidak ditemukan.")

    filters = [
        "pr.user_id = :user_id",
        "(:start_at IS NULL OR pr.created_at >= :start_at)",
        "(:end_at IS NULL OR pr.created_at <= :end_at)",
    ]
    params = {
        "user_id": user_id,
        "start_at": _start_datetime(start_date),
        "end_at": _end_datetime(end_date),
    }

    rows = db.execute(
        text(f"""
            SELECT
                pr.id,
                pr.created_at,
                pr.predicted_class,
                pr.confidence,
                pr.prob_belum_masak,
                pr.prob_masak,
                pr.prob_terlalu_masak,
                pr.input_source,
                pr.image_width,
                pr.image_height,
                pr.file_size_bytes,
                pr.image_processed_url,
                pr.image_thumbnail_url
            FROM prediction_records pr
            WHERE {' AND '.join(filters)}
            ORDER BY pr.created_at ASC
        """),
        params,
    ).fetchall()

    records = [
        {
            "id": str(row[0]),
            "created_at": _excel_safe_datetime(row[1]),
            "predicted_class": row[2],
            "confidence": float(row[3] or 0),
            "prob_belum_masak": float(row[4] or 0),
            "prob_masak": float(row[5] or 0),
            "prob_terlalu_masak": float(row[6] or 0),
            "input_source": row[7] or "-",
            "image_width": row[8],
            "image_height": row[9],
            "file_size_bytes": row[10] or 0,
            "image_processed_url": row[11],
            "image_thumbnail_url": row[12],
        }
        for row in rows
    ]

    return {
        "user": {
            "id": str(user_row[0]),
            "name": user_row[1],
            "phone_number": user_row[2],
            "created_at": _excel_safe_datetime(user_row[3]),
        },
        "records": records,
        "start_date": start_date,
        "end_date": end_date,
    }


def get_admin_prediction_report_data(
    db,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user_id: Optional[str] = None,
    predicted_class: Optional[str] = None,
) -> Dict[str, Any]:
    filters = [
        "(:start_at IS NULL OR pr.created_at >= :start_at)",
        "(:end_at IS NULL OR pr.created_at <= :end_at)",
        "(:user_id IS NULL OR pr.user_id = :user_id)",
        "(:predicted_class IS NULL OR pr.predicted_class = :predicted_class)",
    ]
    params = {
        "start_at": _start_datetime(start_date),
        "end_at": _end_datetime(end_date),
        "user_id": user_id,
        "predicted_class": predicted_class,
    }

    rows = db.execute(
        text(f"""
            SELECT
                pr.id,
                pr.created_at,
                pr.predicted_class,
                pr.confidence,
                pr.prob_belum_masak,
                pr.prob_masak,
                pr.prob_terlalu_masak,
                pr.input_source,
                pr.image_width,
                pr.image_height,
                pr.file_size_bytes,
                pr.image_processed_url,
                pr.image_thumbnail_url,
                u.id,
                u.full_name,
                u.phone_number,
                u.is_active
            FROM prediction_records pr
            LEFT JOIN users u ON u.id = pr.user_id
            WHERE {' AND '.join(filters)}
            ORDER BY pr.created_at ASC
        """),
        params,
    ).fetchall()

    records = [
        {
            "id": str(row[0]),
            "created_at": _excel_safe_datetime(row[1]),
            "predicted_class": row[2],
            "confidence": float(row[3] or 0),
            "prob_belum_masak": float(row[4] or 0),
            "prob_masak": float(row[5] or 0),
            "prob_terlalu_masak": float(row[6] or 0),
            "input_source": row[7] or "-",
            "image_width": row[8],
            "image_height": row[9],
            "file_size_bytes": row[10] or 0,
            "image_processed_url": row[11],
            "image_thumbnail_url": row[12],
            "user_id": str(row[13]) if row[13] else None,
            "user_name": row[14] or "Tidak diketahui",
            "user_phone_number": row[15] or "-",
            "is_active": bool(row[16]) if row[16] is not None else False,
        }
        for row in rows
    ]

    user_count = db.execute(
        text("SELECT COUNT(*) FROM users WHERE role = 'user'")
    ).scalar()

    active_user_count = db.execute(
        text("SELECT COUNT(*) FROM users WHERE role = 'user' AND is_active = TRUE")
    ).scalar()

    return {
        "records": records,
        "start_date": start_date,
        "end_date": end_date,
        "selected_user_id": user_id,
        "selected_class": predicted_class,
        "total_registered_users": int(user_count or 0),
        "total_active_users": int(active_user_count or 0),
    }


def _calculate_summary(records: Iterable[Dict[str, Any]]) -> Dict[str, Any]:
    records = list(records)
    counts = {key: 0 for key in CLASS_LABELS}
    confidence_values: List[float] = []

    for item in records:
        class_name = item.get("predicted_class")
        if class_name in counts:
            counts[class_name] += 1
        confidence_values.append(float(item.get("confidence") or 0))

    average_confidence = (
        sum(confidence_values) / len(confidence_values)
        if confidence_values
        else 0
    )

    return {
        "total": len(records),
        "counts": counts,
        "average_confidence": average_confidence,
    }


def _style_title(ws, title: str, subtitle: Optional[str] = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20


def _apply_table_header(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def _apply_body_borders(ws, start_row: int, end_row: int, columns: int) -> None:
    for row_index in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor="FFFFFF" if row_index % 2 else "F7FAF6")
        for col in range(1, columns + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _auto_width(ws, max_width: int = 38) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        ws.column_dimensions[column_letter].width = min(max(width + 2, 10), max_width)


def _style_kpi_card(ws, cell_range: str, label: str, value: Any, fill_color: str) -> None:
    ws.merge_cells(cell_range)
    start_cell = cell_range.split(":")[0]
    cell = ws[start_cell]
    cell.value = f"{label}\n{value}"
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color="FFFFFF", bold=True, size=13)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="medium", color=fill_color),
        right=Side(style="medium", color=fill_color),
        top=Side(style="medium", color=fill_color),
        bottom=Side(style="medium", color=fill_color),
    )


def _add_distribution_charts(
    ws,
    start_row: int,
    label_col: int,
    value_col: int,
    anchor_bar: str,
    anchor_doughnut: str,
) -> None:
    data = Reference(
        ws,
        min_col=value_col,
        min_row=start_row,
        max_row=start_row + 3,
    )
    categories = Reference(
        ws,
        min_col=label_col,
        min_row=start_row + 1,
        max_row=start_row + 3,
    )

    # Grafik batang memakai tiga seri terpisah agar legenda menampilkan
    # Belum Masak, Masak, dan Terlalu Masak seperti diagram doughnut.
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Jumlah Prediksi per Kelas"

    # Setiap baris tabel distribusi menjadi satu seri. Kolom pertama
    # dipakai sebagai nama seri/legenda, kolom kedua sebagai nilainya.
    bar_data = Reference(
        ws,
        min_col=label_col,
        max_col=value_col,
        min_row=start_row + 1,
        max_row=start_row + 3,
    )
    bar.add_data(
        bar_data,
        titles_from_data=True,
        from_rows=True,
    )

    # Hanya ada satu kelompok nilai; legenda di bawah grafik menjelaskan
    # warna masing-masing kelas.
    bar_categories = Reference(
        ws,
        min_col=value_col,
        min_row=start_row,
        max_row=start_row,
    )
    bar.set_categories(bar_categories)
    bar.height = 7.8
    bar.width = 14.8
    bar.legend.position = "b"
    bar.varyColors = False
    bar.gapWidth = 70
    bar.y_axis.majorGridlines = None
    bar.y_axis.scaling.min = 0
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.showSerName = False
    bar.dataLabels.showCatName = False
    bar.dataLabels.showLegendKey = False
    bar.graphicalProperties = GraphicalProperties(
        noFill=True,
        ln=LineProperties(noFill=True),
    )
    ws.add_chart(bar, anchor_bar)

    # Doughnut hanya menampilkan persentase; nama kelas dibaca dari legenda.
    # Ini menghindari teks panjang seperti "Jumlah, Belum Masak, 3, 37%".
    doughnut = DoughnutChart()
    doughnut.title = "Komposisi Hasil Prediksi"
    doughnut.add_data(data, titles_from_data=True)
    doughnut.set_categories(categories)
    doughnut.height = 7.8
    doughnut.width = 12.8
    doughnut.holeSize = 62
    doughnut.firstSliceAng = 270
    doughnut.varyColors = True
    doughnut.legend.position = "b"
    doughnut.dataLabels = DataLabelList()
    doughnut.dataLabels.showPercent = True
    doughnut.dataLabels.showVal = False
    doughnut.dataLabels.showSerName = False
    doughnut.dataLabels.showCatName = False
    doughnut.dataLabels.showLegendKey = False
    doughnut.dataLabels.showLeaderLines = False
    doughnut.graphicalProperties = GraphicalProperties(
        noFill=True,
        ln=LineProperties(noFill=True),
    )
    ws.add_chart(doughnut, anchor_doughnut)


def _add_daily_trend_sheet(wb: Workbook, records: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Tren Harian")
    _style_title(ws, "Tren Prediksi Harian", "Rekap jumlah hasil klasifikasi berdasarkan tanggal")

    daily: Dict[str, Dict[str, int]] = {}
    for item in records:
        created_at = item.get("created_at")
        if not created_at:
            continue
        day_key = created_at.strftime("%Y-%m-%d")
        daily.setdefault(day_key, {"belum_masak": 0, "masak": 0, "terlalu_masak": 0, "total": 0})
        class_name = item.get("predicted_class")
        if class_name in CLASS_LABELS:
            daily[day_key][class_name] += 1
        daily[day_key]["total"] += 1

    headers = ["Tanggal", "Belum Masak", "Masak", "Terlalu Masak", "Total"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=5, column=col, value=header)
    _apply_table_header(ws, 5, len(headers))

    row_index = 6
    for day_key in sorted(daily):
        values = daily[day_key]
        ws.cell(row=row_index, column=1, value=datetime.strptime(day_key, "%Y-%m-%d"))
        ws.cell(row=row_index, column=2, value=values["belum_masak"])
        ws.cell(row=row_index, column=3, value=values["masak"])
        ws.cell(row=row_index, column=4, value=values["terlalu_masak"])
        ws.cell(row=row_index, column=5, value=values["total"])
        ws.cell(row=row_index, column=6, value=datetime.strptime(day_key, "%Y-%m-%d").strftime("%d %b"))
        ws.cell(row=row_index, column=1).number_format = "dd mmm yyyy"
        row_index += 1

    if row_index == 6:
        ws.merge_cells("A8:E10")
        ws["A8"] = "Belum ada data pada periode yang dipilih."
        ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A8"].fill = ACCENT_FILL
        ws["A8"].font = Font(bold=True, color="666666")
    else:
        _apply_body_borders(ws, 6, row_index - 1, len(headers))
        ws.auto_filter.ref = f"A5:E{row_index - 1}"

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.style = 10
        chart.title = "Komposisi Prediksi per Hari"
        chart.y_axis.majorGridlines = None
        chart.y_axis.scaling.min = 0
        chart.gapWidth = 45
        data = Reference(ws, min_col=2, max_col=4, min_row=5, max_row=row_index - 1)
        categories = Reference(ws, min_col=6, min_row=6, max_row=row_index - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 18
        chart.legend.position = "b"
        # Grafik tren harian tidak memakai label angka.
        # Nilai rinci sudah tersedia pada tabel sehingga grafik tetap bersih.
        chart.dataLabels = None

        # Hilangkan garis/border luar chart dan area plot.
        chart.graphicalProperties = GraphicalProperties(
            noFill=True,
            ln=LineProperties(noFill=True),
        )

        colors = ["70AD47", "ED7D31", "C00000"]
        titles = ["Belum Masak", "Masak", "Terlalu Masak"]
        # Judul seri sudah diambil otomatis dari header tabel melalui
        # chart.add_data(..., titles_from_data=True). Jangan mengisi
        # series.title menggunakan string biasa karena openpyxl 3.1.x
        # mengharuskan objek SeriesLabel.
        for series, color in zip(chart.series, colors):
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color

        ws.add_chart(chart, "G5")

    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].hidden = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale = 90



def _build_daily_summary(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    daily: Dict[str, Dict[str, Any]] = {}

    for item in records:
        created_at = item.get("created_at")
        if not created_at:
            continue

        day_key = created_at.strftime("%Y-%m-%d")
        daily.setdefault(
            day_key,
            {
                "date": created_at.date(),
                "belum_masak": 0,
                "masak": 0,
                "terlalu_masak": 0,
                "total": 0,
                "confidence_sum": 0.0,
            },
        )

        class_name = item.get("predicted_class")
        if class_name in CLASS_LABELS:
            daily[day_key][class_name] += 1

        daily[day_key]["total"] += 1
        daily[day_key]["confidence_sum"] += float(
            item.get("confidence") or 0
        )

    result = []

    for day_key in sorted(daily):
        item = daily[day_key]
        average_confidence = (
            item["confidence_sum"] / item["total"]
            if item["total"]
            else 0
        )

        result.append(
            {
                "date": item["date"],
                "belum_masak": item["belum_masak"],
                "masak": item["masak"],
                "terlalu_masak": item["terlalu_masak"],
                "total": item["total"],
                "average_confidence": average_confidence,
            }
        )

    return result


def _write_daily_summary_section(
    ws,
    records: List[Dict[str, Any]],
    start_row: int,
    title: str = "Rekap Prediksi Harian",
) -> int:
    daily_rows = _build_daily_summary(records)

    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=8,
    )
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = TITLE_FILL
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    ws.row_dimensions[start_row].height = 26

    description_row = start_row + 1
    ws.merge_cells(
        start_row=description_row,
        start_column=1,
        end_row=description_row,
        end_column=8,
    )
    description_cell = ws.cell(
        row=description_row,
        column=1,
        value=(
            "Setiap baris menunjukkan ringkasan seluruh prediksi "
            "yang dilakukan pada satu tanggal."
        ),
    )
    description_cell.font = Font(italic=True, color="666666")
    description_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    header_row = start_row + 3
    headers = [
        "No",
        "Tanggal",
        "Belum Masak",
        "Masak",
        "Terlalu Masak",
        "Total Prediksi",
        "Rata-rata Confidence",
        "Keterangan",
    ]

    for column, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=column, value=header)

    _apply_table_header(ws, header_row, len(headers))

    if not daily_rows:
        empty_row = header_row + 1
        ws.merge_cells(
            start_row=empty_row,
            start_column=1,
            end_row=empty_row,
            end_column=8,
        )
        ws.cell(
            row=empty_row,
            column=1,
            value="Belum ada data prediksi pada periode ini.",
        )
        ws.cell(row=empty_row, column=1).alignment = Alignment(
            horizontal="center"
        )
        return empty_row

    for index, item in enumerate(daily_rows, start=1):
        row = header_row + index
        dominant_class = max(
            ("belum_masak", "masak", "terlalu_masak"),
            key=lambda key: item[key],
        )

        values = [
            index,
            item["date"],
            item["belum_masak"],
            item["masak"],
            item["terlalu_masak"],
            item["total"],
            item["average_confidence"] / 100,
            f"Didominasi {CLASS_LABELS[dominant_class]}",
        ]

        for column, value in enumerate(values, start=1):
            ws.cell(row=row, column=column, value=value)

        ws.cell(row=row, column=2).number_format = "dd mmm yyyy"
        ws.cell(row=row, column=7).number_format = "0.00%"
        ws.row_dimensions[row].height = 24

        fill = PatternFill(
            "solid",
            fgColor="F8FBF6" if index % 2 else "EEF5EA",
        )

        for column in range(1, 9):
            cell = ws.cell(row=row, column=column)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if column != 8 else "left",
                vertical="center",
                wrap_text=True,
            )

    last_row = header_row + len(daily_rows)
    ws.auto_filter.ref = f"A{header_row}:H{last_row}"

    return last_row

def build_user_report_workbook(data: Dict[str, Any]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"

    user = data["user"]
    records = data["records"]
    summary = _calculate_summary(records)
    period_label = build_period_label(data["start_date"], data["end_date"])

    _style_title(
        ws,
        "Laporan Prediksi SawitVision AI",
        "Laporan pribadi hasil klasifikasi kematangan buah kelapa sawit",
    )

    ws.merge_cells("A4:B4")
    ws["A4"] = "Informasi Laporan"
    ws["A4"].fill = HEADER_FILL
    ws["A4"].font = WHITE_FONT
    ws["A4"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Nama Pengguna", user["name"]),
        ("Nomor Telepon", user["phone_number"]),
        ("Periode Laporan", period_label),
        ("Tanggal Dibuat", datetime.now()),
    ]

    for idx, (label, value) in enumerate(info_rows, start=5):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).font = BOLD_FONT
        ws.cell(row=idx, column=1).fill = SUBHEADER_FILL
        ws.cell(row=idx, column=1).border = THIN_BORDER
        ws.cell(row=idx, column=2).border = THIN_BORDER
        ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True)

    ws["B8"].number_format = "dd mmm yyyy hh:mm"

    _style_kpi_card(ws, "D4:F6", "TOTAL PREDIKSI", summary["total"], "2F7D32")
    _style_kpi_card(
        ws,
        "G4:I6",
        "RATA-RATA CONFIDENCE",
        f"{summary['average_confidence']:.2f}%",
        "1F4E78",
    )
    _style_kpi_card(
        ws,
        "D8:F10",
        "HASIL TERBANYAK",
        CLASS_LABELS.get(
            max(summary["counts"], key=summary["counts"].get)
            if summary["total"] else "belum_masak",
            "Belum Ada Data",
        ) if summary["total"] else "Belum Ada Data",
        "C55A11",
    )
    _style_kpi_card(
        ws,
        "G8:I10",
        "DATA TERSIMPAN",
        f"{summary['total']} record",
        "8064A2",
    )

    distribution_start = 13
    ws.cell(row=distribution_start, column=1, value="Kelas")
    ws.cell(row=distribution_start, column=2, value="Jumlah")
    _apply_table_header(ws, distribution_start, 2)

    for offset, class_name in enumerate(CLASS_LABELS, start=1):
        ws.cell(
            row=distribution_start + offset,
            column=1,
            value=CLASS_LABELS[class_name],
        )
        ws.cell(
            row=distribution_start + offset,
            column=2,
            value=summary["counts"][class_name],
        )

    _apply_body_borders(
        ws,
        distribution_start + 1,
        distribution_start + 3,
        2,
    )
    _add_distribution_charts(
        ws,
        distribution_start,
        1,
        2,
        "D13",
        "K13",
    )

    ws_detail = wb.create_sheet("Detail Prediksi")
    _style_title(
        ws_detail,
        "Detail Hasil Prediksi",
        f"Pengguna: {user['name']} | Periode: {period_label}",
    )

    headers = [
        "No",
        "Tanggal dan Waktu",
        "Kelas Prediksi",
        "Confidence",
        "Prob. Belum Masak",
        "Prob. Masak",
        "Prob. Terlalu Masak",
        "Sumber Input",
        "Lebar",
        "Tinggi",
        "Ukuran File (KB)",
        "URL Gambar",
    ]

    header_row = 4
    for col, header in enumerate(headers, start=1):
        ws_detail.cell(row=header_row, column=col, value=header)
    _apply_table_header(ws_detail, header_row, len(headers))

    for index, item in enumerate(records, start=1):
        row = header_row + index
        image_url = (
            item.get("image_processed_url")
            or item.get("image_thumbnail_url")
            or ""
        )
        values = [
            index,
            item.get("created_at"),
            CLASS_LABELS.get(
                item.get("predicted_class"),
                item.get("predicted_class"),
            ),
            float(item.get("confidence") or 0) / 100,
            float(item.get("prob_belum_masak") or 0) / 100,
            float(item.get("prob_masak") or 0) / 100,
            float(item.get("prob_terlalu_masak") or 0) / 100,
            item.get("input_source"),
            item.get("image_width"),
            item.get("image_height"),
            round(float(item.get("file_size_bytes") or 0) / 1024, 2),
            image_url,
        ]

        for col, value in enumerate(values, start=1):
            ws_detail.cell(row=row, column=col, value=value)

        ws_detail.cell(row=row, column=2).number_format = "dd/mm/yyyy hh:mm"
        for col in (4, 5, 6, 7):
            ws_detail.cell(row=row, column=col).number_format = "0.00%"

        if image_url:
            ws_detail.cell(row=row, column=12).hyperlink = image_url
            ws_detail.cell(row=row, column=12).style = "Hyperlink"

    if records:
        _apply_body_borders(
            ws_detail,
            header_row + 1,
            header_row + len(records),
            len(headers),
        )
        ws_detail.auto_filter.ref = (
            f"A{header_row}:L{header_row + len(records)}"
        )

    ws_detail.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30
    for col in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        ws.column_dimensions[col].width = 12
    ws.sheet_view.zoomScale = 85
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    _auto_width(ws_detail)
    ws_detail.sheet_view.zoomScale = 85
    ws_daily = wb.create_sheet("Rekap Harian")
    _style_title(
        ws_daily,
        "Rekap Prediksi Harian",
        f"Pengguna: {user['name']} | Periode: {period_label}",
    )
    _write_daily_summary_section(
        ws_daily,
        records,
        start_row=4,
        title="Ringkasan per Tanggal",
    )
    ws_daily.freeze_panes = "A8"
    ws_daily.column_dimensions["A"].width = 7
    ws_daily.column_dimensions["B"].width = 18
    for column in ["C", "D", "E", "F", "G"]:
        ws_daily.column_dimensions[column].width = 18
    ws_daily.column_dimensions["H"].width = 28
    ws_daily.sheet_view.zoomScale = 90

    # Laporan pengguna cukup tiga sheet: Ringkasan, Detail Prediksi,
    # dan Rekap Harian. Tren Harian sengaja tidak dibuat.

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_admin_report_workbook(data: Dict[str, Any]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    records = data["records"]
    summary = _calculate_summary(records)
    period_label = build_period_label(data["start_date"], data["end_date"])

    _style_title(
        ws,
        "Laporan Global SawitVision AI",
        "Rekap seluruh hasil klasifikasi pengguna",
    )

    ws.merge_cells("A4:B4")
    ws["A4"] = "Informasi & Filter Laporan"
    ws["A4"].fill = HEADER_FILL
    ws["A4"].font = WHITE_FONT
    ws["A4"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Periode Laporan", period_label),
        ("Tanggal Dibuat", datetime.now()),
        (
            "Filter Kelas",
            CLASS_LABELS.get(data["selected_class"], "Semua kelas")
            if data["selected_class"] else "Semua kelas",
        ),
        ("Filter User", data["selected_user_id"] or "Semua user"),
    ]

    for idx, (label, value) in enumerate(info_rows, start=5):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).font = BOLD_FONT
        ws.cell(row=idx, column=1).fill = SUBHEADER_FILL
        ws.cell(row=idx, column=1).border = THIN_BORDER
        ws.cell(row=idx, column=2).border = THIN_BORDER
        ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True)

    ws["B6"].number_format = "dd mmm yyyy hh:mm"

    _style_kpi_card(ws, "D4:F6", "TOTAL USER", data["total_registered_users"], "2F7D32")
    _style_kpi_card(ws, "G4:I6", "USER AKTIF", data["total_active_users"], "548235")
    _style_kpi_card(ws, "J4:L6", "TOTAL PREDIKSI", summary["total"], "C55A11")
    _style_kpi_card(
        ws,
        "D8:F10",
        "RATA-RATA CONFIDENCE",
        f"{summary['average_confidence']:.2f}%",
        "1F4E78",
    )
    _style_kpi_card(
        ws,
        "G8:I10",
        "HASIL MASAK",
        summary["counts"]["masak"],
        "ED7D31",
    )
    

    distribution_start = 13
    ws.cell(row=distribution_start, column=1, value="Kelas")
    ws.cell(row=distribution_start, column=2, value="Jumlah")
    _apply_table_header(ws, distribution_start, 2)

    for offset, class_name in enumerate(CLASS_LABELS, start=1):
        ws.cell(
            row=distribution_start + offset,
            column=1,
            value=CLASS_LABELS[class_name],
        )
        ws.cell(
            row=distribution_start + offset,
            column=2,
            value=summary["counts"][class_name],
        )

    _apply_body_borders(
        ws,
        distribution_start + 1,
        distribution_start + 3,
        2,
    )
    _add_distribution_charts(
        ws,
        distribution_start,
        1,
        2,
        "D13",
        "K13",
    )

    # Rekap per user
    ws_users = wb.create_sheet("Rekap & Harian")
    _style_title(
        ws_users,
        "Rekap Prediksi per Pengguna",
        f"Periode: {period_label}",
    )

    user_summary: Dict[str, Dict[str, Any]] = {}
    for item in records:
        key = item.get("user_id") or item.get("user_phone_number") or "unknown"
        user_summary.setdefault(
            key,
            {
                "name": item.get("user_name") or "Tidak diketahui",
                "phone_number": item.get("user_phone_number") or "-",
                "total": 0,
                "belum_masak": 0,
                "masak": 0,
                "terlalu_masak": 0,
                "confidence_sum": 0.0,
            },
        )
        current = user_summary[key]
        current["total"] += 1
        class_name = item.get("predicted_class")
        if class_name in CLASS_LABELS:
            current[class_name] += 1
        current["confidence_sum"] += float(item.get("confidence") or 0)

    user_headers = [
        "No",
        "Nama",
        "Nomor Telepon",
        "Total Prediksi",
        "Belum Masak",
        "Masak",
        "Terlalu Masak",
        "Rata-rata Confidence",
    ]

    for col, header in enumerate(user_headers, start=1):
        ws_users.cell(row=4, column=col, value=header)
    _apply_table_header(ws_users, 4, len(user_headers))

    for index, item in enumerate(
        sorted(
            user_summary.values(),
            key=lambda value: value["total"],
            reverse=True,
        ),
        start=1,
    ):
        row = 4 + index
        avg_confidence = (
            item["confidence_sum"] / item["total"]
            if item["total"]
            else 0
        )
        values = [
            index,
            item["name"],
            item["phone_number"],
            item["total"],
            item["belum_masak"],
            item["masak"],
            item["terlalu_masak"],
            avg_confidence / 100,
        ]
        for col, value in enumerate(values, start=1):
            ws_users.cell(row=row, column=col, value=value)
        ws_users.cell(row=row, column=8).number_format = "0.00%"

    if user_summary:
        last_row = 4 + len(user_summary)
        _apply_body_borders(ws_users, 5, last_row, len(user_headers))
        ws_users.auto_filter.ref = f"A4:H{last_row}"



    user_table_last_row = 4 + len(user_summary)
    daily_section_start = max(user_table_last_row + 4, 10)

    _write_daily_summary_section(
        ws_users,
        records,
        start_row=daily_section_start,
        title="Rekap Prediksi Harian",
    )

    ws_users.freeze_panes = "A5"
    ws_users.column_dimensions["A"].width = 7
    ws_users.column_dimensions["B"].width = 24
    ws_users.column_dimensions["C"].width = 30
    for column in ["D", "E", "F", "G", "H"]:
        ws_users.column_dimensions[column].width = 18

    # Seluruh prediksi
    ws_all = wb.create_sheet("Seluruh Prediksi")
    _style_title(
        ws_all,
        "Seluruh Data Prediksi",
        (
            f"Periode: {period_label} | "
            "Setiap baris mewakili satu gambar dan satu hasil prediksi."
        ),
    )

    all_headers = [
        "No",
        "ID Prediksi",
        "Gambar Ke-",
        "Nama User",
        "Nomor Telepon",
        "Tanggal dan Waktu",
        "Kelas Prediksi",
        "Confidence Gambar",
        "Prob. Belum Masak",
        "Prob. Masak",
        "Prob. Terlalu Masak",
        "Sumber Input",
        "Ukuran File (KB)",
        "URL Gambar",
    ]

    header_row = 4

    for column, header in enumerate(all_headers, start=1):
        ws_all.cell(row=header_row, column=column, value=header)

    _apply_table_header(ws_all, header_row, len(all_headers))

    user_image_counters: Dict[str, int] = {}

    for index, item in enumerate(records, start=1):
        row = header_row + index
        user_key = (
            item.get("user_id")
            or item.get("user_phone_number")
            or "unknown"
        )

        user_image_counters[user_key] = (
            user_image_counters.get(user_key, 0) + 1
        )

        image_url = (
            item.get("image_processed_url")
            or item.get("image_thumbnail_url")
            or ""
        )

        values = [
            index,
            item.get("id"),
            user_image_counters[user_key],
            item.get("user_name"),
            item.get("user_phone_number"),
            item.get("created_at"),
            CLASS_LABELS.get(
                item.get("predicted_class"),
                item.get("predicted_class"),
            ),
            float(item.get("confidence") or 0) / 100,
            float(item.get("prob_belum_masak") or 0) / 100,
            float(item.get("prob_masak") or 0) / 100,
            float(item.get("prob_terlalu_masak") or 0) / 100,
            item.get("input_source"),
            round(
                float(item.get("file_size_bytes") or 0) / 1024,
                2,
            ),
            image_url,
        ]

        for column, value in enumerate(values, start=1):
            ws_all.cell(row=row, column=column, value=value)

        ws_all.cell(row=row, column=6).number_format = (
            "dd/mm/yyyy hh:mm:ss"
        )

        for column in (8, 9, 10, 11):
            ws_all.cell(row=row, column=column).number_format = (
                "0.00%"
            )

        if image_url:
            ws_all.cell(row=row, column=14).hyperlink = image_url
            ws_all.cell(row=row, column=14).style = "Hyperlink"

        row_fill = PatternFill(
            "solid",
            fgColor="FFFFFF" if index % 2 else "F4F8F1",
        )

        for column in range(1, len(all_headers) + 1):
            cell = ws_all.cell(row=row, column=column)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column in (4, 5, 7, 12, 14)
                    else "center"
                ),
                vertical="center",
                wrap_text=column in (4, 5, 7, 12, 14),
            )

        ws_all.row_dimensions[row].height = 30

    if records:
        last_row = header_row + len(records)
        ws_all.auto_filter.ref = (
            f"A{header_row}:N{last_row}"
        )

    ws_all.freeze_panes = "A5"

    compact_widths = {
        "A": 7,
        "B": 18,
        "C": 11,
        "D": 22,
        "E": 28,
        "F": 21,
        "G": 18,
        "H": 18,
        "I": 19,
        "J": 15,
        "K": 20,
        "L": 15,
        "M": 17,
        "N": 26,
    }

    for column_letter, width in compact_widths.items():
        ws_all.column_dimensions[column_letter].width = width

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30

    for column in [
        "D", "E", "F", "G", "H", "I",
        "J", "K", "L", "M", "N", "O",
        "P", "Q",
    ]:
        ws.column_dimensions[column].width = 12

    ws.sheet_view.zoomScale = 80
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws_users.sheet_view.zoomScale = 90
    ws_all.sheet_view.zoomScale = 85

    # Tambahkan visualisasi tren harian ke laporan admin juga.
    _add_daily_trend_sheet(wb, records)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer